# -*- coding: UTF-8 -*-

import s_autoinstall
os = s_autoinstall.ImportModule("os")
sys = s_autoinstall.ImportModule("sys")
csv = s_autoinstall.ImportModule("csv")
json = s_autoinstall.ImportModule("json")
shutil = s_autoinstall.ImportModule("shutil")
zipfile = s_autoinstall.ImportModule("zipfile")
# openpyxl = s_autoinstall.ImportModule("openpyxl")
import openpyxl
s_log = s_autoinstall.ImportModule("s_log")

# 过滤
DS_Store = ".DS_Store"

def getExePath():
    '''
    获取执行路径
    :return: 路径 str
    '''
    return os.path.dirname(sys.argv[0])

def getFiles(fpath, nodir=False):
    '''
    获取路径下的所有文件路径列表
    :param fpath: 路径 str
    :param nodir: 不包含文件夹 str
    :return: 路径列表 str[]
    '''
    if not os.path.exists(fpath):
        s_log.zError("getFiles:文件路径不存在 | " + str(fpath))
        return [""]

    isdir = os.path.isdir(fpath)
    flist = []

    if (not isdir) or (not nodir):
        flist.append(fpath)
    if isdir:
        fflist = os.listdir(fpath)
        for f in fflist:
            if f.find(DS_Store) == -1:
                flist.extend(getFiles(fpath + '/' + f, nodir))

    return flist

def createDir(fpath):
    '''
    创建文件夹
    :param fpath: 路径 str
    '''
    if not os.path.exists(fpath):
        os.makedirs(fpath)

def removeFile(fpath):
    '''
    移除文件
    :param fpath: 路径 str
    :return: 是否操作成功
    '''
    if not os.path.exists(fpath):
        s_log.zError("removeFile:文件路径不存在 | " + str(fpath))
        return False

    if os.path.isfile(fpath):
        os.remove(fpath)
    elif os.path.isdir(fpath):
        list = os.listdir(fpath)
        for f in list:
            removeFile(fpath + '/' + f)
        os.rmdir(fpath)
    
    return True

def copyFilesTo(fpaths, target):
    '''
    拷贝文件
    :param fpaths: 路径列表 str[]
    :param target: 目标路径 str
    '''
    for f in fpaths:
        if not os.path.exists(f):
            s_log.zError("copyFilesTo:文件路径不存在 | " + str(f))
        else:
            basename = os.path.basename(f)
            cppath = target + '/' + basename

            if os.path.isfile(f):
                shutil.copy(f, cppath)
            elif os.path.isdir(f):
                # 创建文件夹
                createDir(cppath)

                # 获取拷贝列表
                flist = os.listdir(f)
                flist = [(f + '/' + ff)
                         for ff in flist if ff.find(DS_Store) == -1]

                # 遍历拷贝
                copyFilesTo(flist, cppath)

def doZipFile(fpaths, target):
    '''
    压缩文件
    :param fpaths: 路径列表 str[]
    :param target: 目标路径 str
    :return: 是否操作成功 bool
    '''
    try:
        z = zipfile.ZipFile(target, 'w')
        for f in fpaths:
            if not os.path.exists(f):
                s_log.zError("doZipFile:文件路径不存在 | " + str(f))
            else:
                flist = getFiles(f, True)
                for ff in flist:
                    zipname = ff.replace(f, '', 1)
                    z.write(ff, zipname, zipfile.ZIP_DEFLATED)
        z.close()
        return True
    except Exception as err:
        s_log.zError("doZipFile:压缩错误 | " + str(err))
        return False

def doUnzipFile(fpath, target = "./"):
    '''
    解压文件
    :param fpath: 路径 str
    :return: 是否操作成功
    '''
    if not os.path.exists(fpath):
        s_log.zError("doUnzipFile:文件路径不存在 | " + str(fpath))
        return False

    try:
        z = zipfile.ZipFile(fpath, 'r')
        z.extractall(target)
        z.close()
        return True
    except Exception as err:
        s_log.zError("doUnzipFile:解压错误 | " + str(err))
        return False

def loadJSON(fpath):
    '''
    加载json文件
    :param fpath: 路径 str
    :return: 数据 any
    '''
    if not os.path.exists(fpath):
        s_log.zError("loadJson:文件路径不存在 | " + str(fpath))
        return None

    try:
        f = open(fpath, 'r')
        data = json.loads(f.read())
        f.close()
        return data
    except Exception as err:
        s_log.zError("loadJson:加载json失败 | " + str(err))
        return None

def loadXLSX(fpath):
    '''
    加载xlsx文件
    :param fpath: 路径 str
    :return: 数据 {[sheetname]: str[][]}
    '''
    if not os.path.exists(fpath):
        s_log.zError("loadXLSX:文件路径不存在 | " + str(fpath))
        return None

    # 解析sheet
    def parseSheet(ws):
        column, row = ws.max_column, ws.max_row
        data = []
        for r in range(1, row + 1):
            arr = []
            for c in range(1, column + 1):
                cell = ws.cell(r, c)
                if cell != None:
                    v = cell.value
                    arr.append(v)
            data.append(arr)
        return data

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        data = {}
        for i, sheetname in enumerate(wb.sheetnames):
            data.setdefault(sheetname, parseSheet(wb.worksheets[i]))
        wb.close()
        return data
    except Exception as err:
        s_log.zError("loadXLSX:加载xlsx失败 | " + str(err))
        return None

def loadCSV(fpath):
    '''
    加载csv文件
    :param fpath: 路径 str
    :return: 数据 str[][]
    '''
    if not os.path.exists(fpath):
        s_log.zError("loadCSV:文件路径不存在 | " + str(fpath))
        return None

    try:
        f = open(fpath, "r")
        alllines = csv.reader(f)
        data = []
        for item in alllines:
            data.append(item.copy())
        f.close()
        return data
    except Exception as err:
        s_log.zError("loadCSV:加载csv失败 | " + str(err))
        return None

def outputJSON(data, fpath):
    '''
    输出json文件
    :param data: 路径 {} | []
    :param fpath: 路径 str
    :return: 是否操作成功
    '''
    if data == None:
        s_log.zError("outputJSON:过滤无效数据")
        return False

    try:
        sdata = json.dumps(data)
        f = open(fpath, 'w')
        f.write(sdata)
        f.close()
        return True
    except Exception as err:
        s_log.zError("outputJSON:输出json文件错误 | " + str(err))
        return False

def outputXLSX(data, fpath):
    '''
    输出xlsx文件
    :param data: 数据 {[sheetname]: str[][]}
    :param fpath: 路径 str
    :return: 是否操作成功
    '''
    if data == None:
        s_log.zError("outputXLSX:过滤无效数据")
        return False

    try:
        wb = openpyxl.Workbook()
        for _, sheet in enumerate(wb.worksheets):
            wb.remove(sheet)
        for sheetname in data:
            ws = wb.create_sheet(sheetname)
            sheetdate = data[sheetname]
            for rdata in sheetdate:
                ws.append(rdata)
        wb.save(fpath)
        wb.close()
        return True
    except Exception as err:
        s_log.zError("outputXLSX:输出xlsx文件错误 | " + str(err))
        return False

def outputCSV(data, fpath):
    '''
    输出csv文件
    :param data: 数据 str[]
    :param fpath: 路径 str
    :return: 是否操作成功
    '''
    if data == None:
        s_log.zError("outputCSV:过滤无效数据")
        return False

    try:
        f = open(fpath, 'w')
        csvcontent = ""
        for s in data:
            csvcontent += s + "\n"
        f.write(csvcontent)
        f.close()
        return True
    except Exception as err:
        s_log.zError("outputCSV:输出csv文件错误 | " + str(err))
        return False
